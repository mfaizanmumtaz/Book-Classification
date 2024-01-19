from langchain.schema.runnable import RunnablePassthrough,RunnableParallel
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.schema.output_parser import StrOutputParser
from langchain.prompts import ChatPromptTemplate
from openpyxl.utils import get_column_letter
from langchain.prompts import PromptTemplate
from langchain_openai import ChatOpenAI
from openpyxl.styles import Alignment
from collections import defaultdict
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import os
from dotenv import load_dotenv
load_dotenv()

os.environ["LANGCHAIN_TRACING_V2"] = "true"
os.environ["LANGCHAIN_ENDPOINT"] = "https://api.smith.langchain.com"
os.environ["LANGCHAIN_PROJECT"] = "Books Classifications"

OpenAI = ChatOpenAI(model="gpt-3.5-turbo-1106",temperature=0.4).with_fallbacks([ChatOpenAI(model="gpt-3.5-turbo-1106",temperature=0.4)])

def pack_to_excel(response, excel_data):
    BooksNames = excel_data['Book'].values

    Summary = "\n\n".join(response["Summary"])
    SWEBOK_Area_Category = "\n\n".join(response["SWEBOK_Area_Category"])
    Primary_SWEBOK_Area_Percentage = response["Primary_SWEBOK_Area_Percentage"]

    expanded_BookNames = []
    expanded_Summary = []
    expanded_SWEBOK_Area_Category = []
    expanded_Primary_SWEBOK_Area_Percentage = []

    for i, percentage in enumerate(Primary_SWEBOK_Area_Percentage):
        expanded_BookNames.append(BooksNames[i % len(BooksNames)])
        expanded_Summary.append(Summary)
        expanded_SWEBOK_Area_Category.append(SWEBOK_Area_Category)
        expanded_Primary_SWEBOK_Area_Percentage.append(percentage)

    output_data = pd.DataFrame({
        'BookNames': expanded_BookNames,
        'Summary': expanded_Summary,
        'SWEBOK_Area_Category': expanded_SWEBOK_Area_Category,
        'Primary_SWEBOK_Area_Percentage': expanded_Primary_SWEBOK_Area_Percentage
    })

    output_file_path = 'output_data.xlsx'
    output_data.to_excel(output_file_path, index=False)

    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

        # Adjusting row heights based on content
        for row in worksheet.iter_rows():
            max_line_count = 1
            for cell in row:
                if cell.value:
                    line_count = str(cell.value).count('\n') + 1
                    if line_count > max_line_count:
                        max_line_count = line_count
            # Approximate adjustment, you might need to fine-tune the multiplication factor
            worksheet.row_dimensions[cell.row].height = max_line_count * 15  

    workbook.save(output_file_path)
    st.info("Output Data from the first some rows:")
    st.dataframe(output_data[:5])

    st.download_button(
        label="Download Excel file",
        data=open(output_file_path, "rb"),
        file_name="output_data.xlsx",
        mime="application/vnd.ms-excel"
    )

SummaryPromptPrompt = PromptTemplate.from_template("""Your task is to generate a summary for each section of the book '{Book}', specifically from chapter '{Chapter}'. These sections are delimited with ```. The sections belong to chapter '{Chapter}' of the book '{Book}'. Emphasize the key points and insights of each section. Present your output by starting with the chapter name as the title, followed by a concise introduction to the chapter, and then a summary of each section. For clarity, use the section names as titles for their respective summaries. Please ensure high-quality work as this is crucial to my career.

----------------

Sections: ```{Sections}```""")

SWEBOK_Area_CategoryChainChain = PromptTemplate.from_template("""Your task is to classify a particular book chapter section into the SWEBOK knowledge areas. For the book '{Book}', chapter '{Chapter}', and section '{Section}', please categorize the section into the most relevant SWEBOK knowledge areas.
Analyze the SWEBOK knowledge areas and determine which of the following SWEBOK knowledge areas the section closely aligns with. The SWEBOK knowledge areas are delimited with ```.
Output should only contain the SWEBOK knowledge areas, separated by commas.
Providing extra output in your response can have adverse effects.
Please do your best it is very important to my career.
                                                              
----------------
                                                              
```
- Software Requirements
- Software Design
- Software Construction
- Software Testing
- Software Maintenance
- Software Configuration Management
- Software Engineering Management
- Software Engineering Process
- Software Engineering Models and Methods
- Software Quality
- Software Engineering Professional Practice
- Software Engineering Economics
- Computing Foundations
- Mathematical Foundations
- Engineering Foundations ```

Your expertise in this classification is vital for the accurate categorization of this section.""")

Primary_SWEBOK_Area_PercentagePrompt = ChatPromptTemplate.from_template("""Please identify the SWEBOK knowledge areas that the book '{Book}', particularly chapter '{Chapter}', primarily focuses on. Additionally, determine the secondary SWEBOK areas associated with this chapter. Provide the percentage prominence for each of these primary and secondary areas, along with a justification for your assessment.Please do your best it is very important to my career.""")

SummaryPromptChain = SummaryPromptPrompt | OpenAI | StrOutputParser()
SWEBOK_Area_CategoryChain = SWEBOK_Area_CategoryChainChain | OpenAI | StrOutputParser()
Primary_SWEBOK_Area_PercentageChain = Primary_SWEBOK_Area_PercentagePrompt | OpenAI | StrOutputParser()

def chunking(data):

    lst_data = []

    current_book = None
    current_chapter = None

    for inex,row in data.iterrows():
        if row["Type"] == "Head":
            current_book = row["Book"]
            current_chapter = row["Titles"]
            
        elif row["Type"] == "Section" and current_book and current_chapter:
            section = row['Titles']
            data = {
                "Book":current_book,
                "Chapter":current_chapter,
                "Section":section
            }
            lst_data.append(data)

    return lst_data[:200]

def merge_sections(data):
    merged_data = defaultdict(lambda: defaultdict(list))

    for entry in data:
        merged_data[entry['Book']][entry['Chapter']].append(entry['Section'])

    list_of_dicts = [
        {'Book': book, 'Chapter': chapter, 'Sections': "\n".join(sections)}
        for book, chapters in merged_data.items()
        for chapter, sections in chapters.items()
    ]
    return list_of_dicts

main_chain = RunnableParallel(
Summary = RunnablePassthrough() | merge_sections | SummaryPromptChain.map(),
SWEBOK_Area_Category = RunnablePassthrough() | SWEBOK_Area_CategoryChain.map(),
Primary_SWEBOK_Area_Percentage = RunnablePassthrough() | Primary_SWEBOK_Area_PercentageChain.map())

st.set_page_config(page_title="Books Classifications", page_icon=":books:", layout="wide")

st.title("Books Classifications")

example_data = {
    'Book': ['Clean Code', 'Clean Code', 'Clean Code', 'Clean Code', 'Clean Code'],
    'Titles': [
        'Chapter 1: Clean Code',
        'There Will Be Code',
        'Bad Code',
        'The Total Cost of Owning a Mess',
        'The Grand Redesign in the Sky'
    ],
    'Chapter': ['Chapter 1', 'Chapter 1', 'Chapter 1', 'Chapter 1', 'Chapter 1'],
    'Type': ['Head', 'Section', 'Section', 'Section', 'Section']
}


example_data = pd.DataFrame(example_data)

st.info("Please Make Sure Your Excel File Format Should Look Like This: ")

st.write(example_data)

file_name = st.file_uploader("Upload your excel file", type=["xlsx"])
if st.button("Submit"):
    if file_name is not None:
        excel_data = pd.read_excel(file_name)

        try:
            data = chunking(excel_data)

            try:
                with st.spinner("Please wait while we process your data..."):
                    st.session_state.response = main_chain.invoke(data)

                st.success("Data Processed Successfully")

                pack_to_excel(st.session_state.response, excel_data)

            except Exception as e:
                st.error(f"Something went wrong, please try again. {e}")

        except Exception as e:
            st.warning("Incorrect File Format! Please Make Sure Your Excel File Format Should Look Like This with 200 rows: ")
            st.dataframe(example_data)